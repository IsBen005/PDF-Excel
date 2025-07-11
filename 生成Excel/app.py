import os
import tabula
import pandas as pd
import PyPDF2
from flask import Flask, request, render_template, jsonify, send_file
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 全局变量用于存储进度信息
progress_01 = {
    'step': 0,
    'total_steps': 0,
    'status': 'Waiting for upload'
}

progress_02 = {
    'step': 0,
    'total_steps': 0,
    'status': 'Waiting for upload'
}


def pdf_to_excel(pdf_path, output_excel_path, start_page, page_interval):
    global progress_01
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        total_pages = len(reader.pages)

    page_range = range(start_page, total_pages + 1)
    total_pages_to_process = len([i for i in range(len(page_range)) if (i - 1) % page_interval == 0])
    total_tasks = total_pages_to_process + 1
    progress_01['total_steps'] = total_tasks
    progress_01['step'] = 0
    progress_01['status'] = 'Processing PDF'

    wb = Workbook()
    wb.remove(wb.active)

    for i, page in enumerate(page_range, start=1):
        if (i - 1) % page_interval == 0:
            page_df = tabula.read_pdf(pdf_path, pages=str(page))
            if page_df:
                combined_df = pd.concat(page_df, ignore_index=True)
                sheet_name = f'Sheet_{(i - 1) // page_interval + 1}'
                ws = wb.create_sheet(sheet_name)
                for col_idx, col in enumerate(combined_df.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col)
                for row_idx, row in combined_df.iterrows():
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx + 2, column=col_idx, value=value)
            progress_01['step'] += 1

    wb.save(output_excel_path)
    progress_01['step'] += 1
    progress_01['status'] = 'Processing new_ben_01.xlsx completed'



def create_new_excel_with_selected_headers(input_excel_path, output_excel_path, selected_cells):
    global progress_02
    wb_input = load_workbook(input_excel_path)
    ws_input = wb_input.active

    total_tasks = len(selected_cells) + 1
    progress_02['total_steps'] = total_tasks
    progress_02['step'] = 0
    progress_02['status'] = 'Creating new_ben_02.xlsx headers'

    wb_output = Workbook()
    ws_output = wb_output.active

    for col_idx, cell in enumerate(selected_cells, start=1):
        value = ws_input[cell].value
        ws_output.cell(row=1, column=col_idx, value=value)
        progress_02['step'] += 1

    wb_output.save(output_excel_path)
    progress_02['step'] += 1


def add_data_to_new_excel(input_excel_path, output_excel_path, data_cells):
    global progress_02
    wb_input = load_workbook(input_excel_path)
    wb_output = load_workbook(output_excel_path)
    ws_output = wb_output.active

    total_sheets = len(wb_input.sheetnames)
    total_tasks = total_sheets * len(data_cells) + 1
    progress_02['total_steps'] += total_tasks
    progress_02['status'] = 'Adding data to new_ben_02.xlsx'

    actual_row = 2
    for sheet_name in wb_input.sheetnames:
        ws_input = wb_input[sheet_name]
        # 获取第一个和第二个数据单元格的值
        first_cell_value = ws_input[data_cells[0]].value if data_cells else None
        second_cell_value = ws_input[data_cells[1]].value if len(data_cells) > 1 else None

        # 检查第一个和第二个数据单元格的值是否满足过滤条件
        if (first_cell_value is None or (isinstance(first_cell_value, str) and len(first_cell_value) > 15)) or \
                (second_cell_value is None or (isinstance(second_cell_value, str) and len(second_cell_value) > 50)):
            continue

        for col_idx, cell in enumerate(data_cells, start=1):
            try:
                value = ws_input[cell].value
                if isinstance(value, str):
                    value = value.encode('utf-8', errors='ignore').decode('utf-8')
                value = value if value else "-"
                ws_output.cell(row=actual_row, column=col_idx, value=value)
            except Exception as e:
                print(f"Error processing cell {cell} in sheet {sheet_name}: {e}")
            progress_02['step'] += 1
        actual_row += 1

    wb_output.save(output_excel_path)
    progress_02['step'] += 1
    progress_02['status'] = 'Processing new_ben_02.xlsx completed'


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        start_page = int(request.form.get('start_page'))
        page_interval = int(request.form.get('page_interval'))
        if file:
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(pdf_path)
            output_excel_path = 'new_ben_01.xlsx'
            pdf_to_excel(pdf_path, output_excel_path, start_page, page_interval)
            return render_template('select_cells.html', excel_path=output_excel_path)
    return render_template('upload.html')



@app.route('/generate_02', methods=['POST'])
def generate_02():
    input_excel_path = 'new_ben_01.xlsx'
    output_excel_path = 'new_ben_02.xlsx'
    selected_cells = request.form.getlist('selected_cells')
    data_cells = request.form.getlist('data_cells')
    create_new_excel_with_selected_headers(input_excel_path, output_excel_path, selected_cells)
    add_data_to_new_excel(input_excel_path, output_excel_path, data_cells)

    # 删除上传的 PDF 和 new_ben_01.xlsx 文件
    for root, dirs, files in os.walk(UPLOAD_FOLDER):
        for file in files:
            if file.endswith('.pdf'):
                os.remove(os.path.join(root, file))
    if os.path.exists('new_ben_01.xlsx'):
        os.remove('new_ben_01.xlsx')

    send_file(output_excel_path, as_attachment=True)

    return "Excel 转换成功！"


@app.route('/progress_01', methods=['GET'])
def get_progress_01():
    global progress_01
    return jsonify(progress_01)


@app.route('/progress_02', methods=['GET'])
def get_progress_02():
    global progress_02
    return jsonify(progress_02)


@app.route('/get_excel_data', methods=['GET'])
def get_excel_data():
    excel_path = request.args.get('path')
    wb = load_workbook(excel_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return jsonify(data)




if __name__ == '__main__':
    app.run(debug=True)
    